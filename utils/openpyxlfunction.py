import openpyxl


def getRowCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_row


def getColCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_column


def readData(file, sheetName, rowno, colno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(rowno, colno).value


def writeData(file, sheetName, rowno, colno, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(rowno, colno).value = data
    workbook.save(file)


# this function will read all the value of a row from the specific column
def readallsheetdata(file, sheetname, colno):
    wb = openpyxl.load_workbook(file)
    totalrow = getRowCount(file, sheetname)
    list = []
    for i in range(1, totalrow + 1):
        data = readData(file, sheetname, i, colno)
        if not data == None:
            list.append(data)
    return list


def totalsheetcount(file):
    wb = openpyxl.load_workbook(file)
    res = len(wb.sheetnames)
    return res


def allsheetName(file):
    wb = openpyxl.load_workbook(file)
    data = wb.sheetnames
    return data


# this function will write a single row with given data in an array.
def writecol(file, sheetname, row, data):
    wb = openpyxl.load_workbook(file)
    sheet = wb[sheetname]
    maxdata = len(data)
    for colnumber in range(1, maxdata + 1):
        sheet.cell(row, colnumber).value = data[colnumber - 1]
        wb.save(file)


# This function will write a column with automatic row detecting technique :p :p
def writecolautomatic(file, sheetname, data):
    wb = openpyxl.load_workbook(file)
    sheet = wb[sheetname]
    maxdata = len(data)
    row= getRowCount(file, sheetname)+1
    for colnumber in range(1, maxdata + 1):
        sheet.cell(row, colnumber).value = data[colnumber - 1]
        wb.save(file)

def readWrite(file_path):
    fileToRead = open(file_path, 'r+')
    data = int(fileToRead.read())
    # print(type(data))
    newdata = data + 1
    fileToRead.seek(0)
    fileToRead.write(str(newdata))
    fileToRead.truncate()
    fileToRead.close()
    return data

# functions for main
